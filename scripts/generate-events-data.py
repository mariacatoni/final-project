#!/usr/bin/env python3
"""
Read `Concert Data.xlsx` (sheet "Concerts by artist") and write `js/events-data.js`.

Columns used: Date, Artist, Venue, City, State, Setlist (optional).
sortKey is UTC midnight ms for the calendar date (matches prior generated data).

Run from repo root:
  python3 scripts/generate-events-data.py
"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
# Rich-detail journal RTFs (not served by the site): ROOT / "unused-files" / "assets" / "<YYYY-MM-DD>" /
XLSX_PATH = ROOT / "Concert Data.xlsx"
OUT_PATH = ROOT / "js" / "events-data.js"
SHEET_NAME = "Concerts by artist"


def cell_date_to_iso(val) -> str:
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, str):
        s = val.strip()
        return s[:10] if len(s) >= 10 else s
    raise ValueError(f"Unexpected date cell: {val!r}")


def sort_key_ms(iso_date: str) -> int:
    y, m, d = map(int, iso_date.split("-"))
    dt = datetime(y, m, d, tzinfo=timezone.utc)
    return int(dt.timestamp() * 1000)


def main() -> int:
    try:
        import openpyxl
    except ImportError:
        print("Missing openpyxl. Install with: pip install openpyxl", file=sys.stderr)
        return 1

    if not XLSX_PATH.is_file():
        print(f"Missing workbook: {XLSX_PATH}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Sheet {SHEET_NAME!r} not found. Available: {wb.sheetnames}", file=sys.stderr)
        return 1

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Workbook has no rows.", file=sys.stderr)
        return 1

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    try:
        ci = {name: header.index(name) for name in ("Date", "Artist", "City", "State", "Venue", "Setlist")}
    except ValueError as e:
        print(f"Missing expected column in header {header!r}: {e}", file=sys.stderr)
        return 1

    events: list[dict] = []
    for row in rows[1:]:
        if row is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        iso = cell_date_to_iso(row[ci["Date"]])
        artist = (row[ci["Artist"]] or "").strip() if row[ci["Artist"]] is not None else ""
        venue = (row[ci["Venue"]] or "").strip() if row[ci["Venue"]] is not None else ""
        city = (row[ci["City"]] or "").strip() if row[ci["City"]] is not None else ""
        state = (row[ci["State"]] or "").strip() if row[ci["State"]] is not None else ""
        raw_setlist = row[ci["Setlist"]]
        setlist = raw_setlist.strip() if isinstance(raw_setlist, str) and raw_setlist.strip() else None

        year = int(iso[:4])
        obj: dict = {
            "date": iso,
            "year": year,
            "artist": artist,
            "venue": venue,
            "city": city,
            "state": state,
            "sortKey": sort_key_ms(iso),
        }
        if setlist:
            obj["setlist"] = setlist
        events.append(obj)

    chunks = [
        "// AUTO-GENERATED from `Concert Data.xlsx`.",
        "// If you update the spreadsheet, run:  python3 scripts/generate-events-data.py",
        "",
        "export const CONCERT_EVENTS = [",
    ]
    for i, obj in enumerate(events):
        dumped = json.dumps(obj, indent=2, ensure_ascii=False)
        block = "  " + dumped.replace("\n", "\n  ")
        if i < len(events) - 1:
            block += ","
        chunks.append(block)
    chunks.append("];")
    chunks.append("")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text("\n".join(chunks), encoding="utf-8")
    print(f"Wrote {len(events)} events to {OUT_PATH.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
